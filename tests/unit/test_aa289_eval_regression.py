"""AA-289 Part B — services/eval/regression.py: regression detection + golden-tours parsing.

No live DB / no live Bedrock / no live S3: _detect_regression is pure, _download_golden_tours
is exercised against a real in-memory openpyxl workbook (same shape as the real fixture) so a
column-name drift in the fixture would actually be caught, not just mocked away.
"""
import asyncio
import io
import json
from decimal import Decimal
from unittest.mock import AsyncMock, MagicMock, patch

import openpyxl

from services.eval.regression import (
    S1_OLD_REGRESSION_THRESHOLD, _detect_regression, _detect_itinerary_regression,
    _download_golden_tours, _get_baseline, _itinerary_compression_summary,
    _load_itinerary_baseline, _write_itinerary_baseline,
)


# ── _detect_regression: s1_old (avg_quality_score) ──────────────────────────

def test_detect_regression_s1_old_no_baseline_is_never_a_regression():
    current = {"avg_quality_score": 3.0}  # even a terrible score, first-ever run
    assert _detect_regression("s1_old", current, None) is False


def test_detect_regression_s1_old_small_drop_not_flagged():
    current = {"avg_quality_score": 8.0}
    baseline = {"avg_quality_score": 8.5}  # 0.5 drop, under threshold
    assert _detect_regression("s1_old", current, baseline) is False


def test_detect_regression_s1_old_big_drop_flagged():
    current = {"avg_quality_score": 6.5}
    baseline = {"avg_quality_score": 8.5}  # 2.0 drop, over the 1.0 threshold
    assert _detect_regression("s1_old", current, baseline) is True


def test_detect_regression_s1_old_drop_exactly_at_threshold_not_flagged():
    """Strictly greater than the threshold, not >= — a drop of exactly 1.0 is noise-adjacent,
    not yet a confirmed regression."""
    current = {"avg_quality_score": 7.5}
    baseline = {"avg_quality_score": 8.5}
    assert (baseline["avg_quality_score"] - current["avg_quality_score"]) == S1_OLD_REGRESSION_THRESHOLD
    assert _detect_regression("s1_old", current, baseline) is False


def test_detect_regression_s1_old_score_improved_not_flagged():
    current = {"avg_quality_score": 9.5}
    baseline = {"avg_quality_score": 8.0}
    assert _detect_regression("s1_old", current, baseline) is False


def test_detect_regression_s1_old_missing_scores_not_flagged():
    """No scored tours in the current run (e.g. all failed to parse) -> nothing to compare,
    must not crash on None - None."""
    current = {"avg_quality_score": None}
    baseline = {"avg_quality_score": 8.0}
    assert _detect_regression("s1_old", current, baseline) is False


# ── _detect_regression: s1_from_atom (gate pass rate) ────────────────────────

def test_detect_regression_s1_from_atom_no_baseline_is_never_a_regression():
    current = {"gate_pass_count": 2}
    assert _detect_regression("s1_from_atom", current, None) is False


def test_detect_regression_s1_from_atom_same_pass_count_not_flagged():
    current = {"gate_pass_count": 4}
    baseline = {"gate_pass_count": 4}
    assert _detect_regression("s1_from_atom", current, baseline) is False


def test_detect_regression_s1_from_atom_fewer_passes_flagged():
    """A tour that used to clear the grounding gate now failing it is a hard correctness
    regression, not noise — any drop counts, unlike s1_old's numeric threshold."""
    current = {"gate_pass_count": 3}
    baseline = {"gate_pass_count": 4}
    assert _detect_regression("s1_from_atom", current, baseline) is True


def test_detect_regression_s1_from_atom_more_passes_not_flagged():
    current = {"gate_pass_count": 4}
    baseline = {"gate_pass_count": 3}
    assert _detect_regression("s1_from_atom", current, baseline) is False


# ── _download_golden_tours: real openpyxl parsing, no mocked DataFrame ──────

def _build_fixture_bytes() -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Golden Tours"
    header = [
        "tour_id", "country", "name", "subtitle", "trip_type", "duration", "group_size",
        "price_usd", "summary", "highlights", "itinerary_summary", "inclusions",
        "best_time_to_go", "expected_quality_score_min", "expected_quality_score_max",
        "expected_failure_codes", "annotation_notes", "chromadb_tags",
    ]
    ws.append(header)
    ws.append([
        "GT-TH-001", "Thailand", "Northern Highlands Traverse", "11 Days | Chiang Mai",
        "trekking", "11 days / 10 nights", "2-8 private", "$3,400",
        "A sustained traverse through northern Thailand's highlands.",
        "Guided ridge walk above Doi Mae Salong\nEvening with a Lisu weaver in Ban Rak Thai",
        "D1: Arrive Chiang Mai. D2-3: Doi Suthep foothills.",
        "Private guiding | Boutique lodges", "November-February", 7.5, 9,
        "None expected", "Strong golden example.", "trekking | thailand",
    ])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_download_golden_tours_parses_real_workbook_shape():
    fixture_bytes = _build_fixture_bytes()
    fake_s3 = MagicMock()
    fake_s3.get_object.return_value = {"Body": io.BytesIO(fixture_bytes)}

    with patch("services.eval.regression.boto3.client", return_value=fake_s3):
        tours = _download_golden_tours()

    assert len(tours) == 1
    t = tours[0]
    assert t["name"] == "Northern Highlands Traverse"
    assert t["country"] == "Thailand"
    assert t["itineraries"] == "D1: Arrive Chiang Mai. D2-3: Doi Suthep foothills."
    # highlights column is \n-joined in the fixture -> must become a real list, not a raw string
    assert t["highlights"] == [
        "Guided ridge walk above Doi Mae Salong",
        "Evening with a Lisu weaver in Ban Rak Thai",
    ]
    assert t["description"] == ""  # no equivalent column in the golden fixture


def test_download_golden_tours_skips_blank_trailing_rows():
    """openpyxl's iter_rows can yield fully-empty rows past the real data (common after manual
    Excel edits/saves) — a row with no tour_id must be skipped, not turned into a fake blank
    tour that would silently drag down avg_quality_score."""
    fixture_bytes = _build_fixture_bytes()
    wb = openpyxl.load_workbook(io.BytesIO(fixture_bytes))
    ws = wb["Golden Tours"]
    ws.append([None] * 18)
    buf = io.BytesIO()
    wb.save(buf)

    fake_s3 = MagicMock()
    fake_s3.get_object.return_value = {"Body": io.BytesIO(buf.getvalue())}
    with patch("services.eval.regression.boto3.client", return_value=fake_s3):
        tours = _download_golden_tours()

    assert len(tours) == 1


# ── AA-353: _itinerary_compression_summary — bucketing + worst-ratio ────────

def _tour_result(day_ratios):
    return {"itinerary_day_ratios": day_ratios}


def test_itinerary_compression_summary_buckets_by_day_count():
    """A 10-day tour goes in "long" (>= ITINERARY_LONG_TOUR_DAY_THRESHOLD=9), a 3-day tour in
    "other" — worst_ratio per bucket is the MIN ratio (most compressed), not an average."""
    long_tour = _tour_result([{"ratio": r} for r in [0.9, 0.5, 1.1]] + [{"ratio": 0.8}] * 7)  # 10 days
    other_tour = _tour_result([{"ratio": 0.95}, {"ratio": 1.05}, {"ratio": 0.85}])  # 3 days

    summary = _itinerary_compression_summary([long_tour, other_tour])

    assert summary["long_tour_count"] == 1
    assert summary["long_worst_ratio"] == 0.5
    assert summary["other_tour_count"] == 1
    assert summary["other_worst_ratio"] == 0.85
    assert summary["total_day_count"] == 13


def test_itinerary_compression_summary_empty_bucket_is_none_not_zero():
    """No long tour in the sample -> long_worst_ratio must be None, not 0 or 1.0 — a silent 0/1.0
    would be misread as either total failure or perfect compliance."""
    other_tour = _tour_result([{"ratio": 0.9}])
    summary = _itinerary_compression_summary([other_tour])
    assert summary["long_worst_ratio"] is None
    assert summary["long_tour_count"] == 0


def test_itinerary_compression_summary_nudge_rate():
    tour = _tour_result([
        {"ratio": 0.9, "nudged": False}, {"ratio": 0.3, "nudged": True},
        {"ratio": 1.0, "nudged": False}, {"ratio": 0.2, "nudged": True},
    ])
    summary = _itinerary_compression_summary([tour])
    assert summary["nudged_day_count"] == 2
    assert summary["total_day_count"] == 4
    assert summary["nudge_rate"] == 0.5


def test_itinerary_compression_summary_ignores_tours_with_no_ratios():
    """A tour whose itinerary wasn't the structured-array contract (empty itinerary_day_ratios,
    e.g. legacy string fallback) must not be counted at all — not as a 0-day tour, not as a
    spurious perfect-ratio tour."""
    summary = _itinerary_compression_summary([_tour_result([]), {}])
    assert summary["long_tour_count"] == 0
    assert summary["other_tour_count"] == 0
    assert summary["total_day_count"] == 0
    assert summary["nudge_rate"] is None


# ── AA-353: baseline file read/write/compare ─────────────────────────────────

def test_write_then_load_itinerary_baseline_roundtrip(tmp_path):
    baseline_path = tmp_path / "itinerary_compression_baseline.json"
    summary = {"long_worst_ratio": 0.55, "other_worst_ratio": 0.7}

    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(baseline_path)):
        _write_itinerary_baseline(summary)
        loaded = _load_itinerary_baseline()

    assert loaded == summary
    assert json.loads(baseline_path.read_text()) == summary


def test_load_itinerary_baseline_missing_file_returns_none(tmp_path):
    missing_path = tmp_path / "does_not_exist.json"
    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(missing_path)):
        assert _load_itinerary_baseline() is None


def test_detect_itinerary_regression_no_baseline_never_flags(tmp_path):
    missing_path = tmp_path / "does_not_exist.json"
    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(missing_path)):
        assert _detect_itinerary_regression({"long_worst_ratio": 0.2}) is False


def test_detect_itinerary_regression_worse_than_baseline_flags(tmp_path):
    baseline_path = tmp_path / "b.json"
    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(baseline_path)):
        _write_itinerary_baseline({"long_worst_ratio": 0.6, "other_worst_ratio": 0.7})
        assert _detect_itinerary_regression({"long_worst_ratio": 0.5, "other_worst_ratio": 0.7}) is True


def test_detect_itinerary_regression_improved_not_flagged(tmp_path):
    baseline_path = tmp_path / "b.json"
    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(baseline_path)):
        _write_itinerary_baseline({"long_worst_ratio": 0.4, "other_worst_ratio": 0.5})
        assert _detect_itinerary_regression({"long_worst_ratio": 0.65, "other_worst_ratio": 0.7}) is False


# ── _get_baseline: Decimal -> float normalization (crash repro) ─────────────

def test_get_baseline_converts_decimal_avg_quality_score_to_float():
    """Reproduces a real crash hit during AA-353's live VERIFY: Postgres NUMERIC comes back as
    decimal.Decimal via asyncpg, while every current-run avg_quality_score is a plain float
    (round(sum(...)/len(...), 3)) — mixing them in _detect_regression's subtraction raised
    TypeError on the actual 20-tour golden-dataset run once a real baseline row existed.
    _get_baseline must hand back a float, not a Decimal, so this can never recur."""
    fake_conn = MagicMock()
    fake_conn.fetchrow = AsyncMock(return_value={
        "prompt_version": "abc123", "avg_quality_score": Decimal("9.900"),
        "avg_words_per_citation": Decimal("14.500"), "gate_pass_count": 4,
    })
    baseline = asyncio.run(_get_baseline(fake_conn, "s1_old", "def456"))
    assert isinstance(baseline["avg_quality_score"], float)
    assert baseline["avg_quality_score"] == 9.9
    assert isinstance(baseline["avg_words_per_citation"], float)


def test_get_baseline_returns_none_when_no_prior_row():
    fake_conn = MagicMock()
    fake_conn.fetchrow = AsyncMock(return_value=None)
    baseline = asyncio.run(_get_baseline(fake_conn, "s1_old", "def456"))
    assert baseline is None


def test_get_baseline_output_feeds_detect_regression_without_crash():
    """End-to-end repro: the exact TypeError seen live only fires when a Decimal baseline meets
    a float current score inside _detect_regression's subtraction — this proves the fixed
    _get_baseline output no longer triggers it, using the same threshold-crossing values that
    would previously have raised."""
    fake_conn = MagicMock()
    fake_conn.fetchrow = AsyncMock(return_value={
        "prompt_version": "abc123", "avg_quality_score": Decimal("8.500"),
        "avg_words_per_citation": None, "gate_pass_count": None,
    })
    baseline = asyncio.run(_get_baseline(fake_conn, "s1_old", "def456"))
    current = {"avg_quality_score": 6.5}  # 2.0 drop, over threshold
    assert _detect_regression("s1_old", current, baseline) is True


def test_detect_itinerary_regression_none_bucket_in_either_run_skipped(tmp_path):
    """A bucket with no tours in EITHER the baseline or the current run (None vs None, or
    None vs a real number) must not trip the regression check — e.g. the golden-20 fixture may
    have zero "long" tours in a given run (AA-339's own caveat, see implementation notes)."""
    baseline_path = tmp_path / "b.json"
    with patch("services.eval.regression.ITINERARY_BASELINE_PATH", str(baseline_path)):
        _write_itinerary_baseline({"long_worst_ratio": None, "other_worst_ratio": 0.7})
        assert _detect_itinerary_regression({"long_worst_ratio": 0.1, "other_worst_ratio": 0.75}) is False
