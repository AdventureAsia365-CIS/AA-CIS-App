"""
services/eval/regression.py — AA-289 Part B: on-demand prompt-regression eval gate.

Runs the REAL pipelines (v1_pipeline._rewrite_tour for S1-old,
content_generation.s1_from_atom.generate_s1_from_atom for S1-from-atom) against a fixed
tour set, tags the result with prompt_version, and compares against the most recent PRIOR
shared.prompt_eval_runs row for that pipeline with a DIFFERENT prompt_version (its baseline).
Writes its own result row either way — the first-ever run for a pipeline has nothing to
compare against yet and just establishes a baseline (not a failure).

ON-DEMAND ONLY (AA-289 explicit constraint): this module has no scheduler of its own. It is
invoked either directly (python -m services.eval.regression) via ECS exec, or by
.github/workflows/eval-regression.yml, which is workflow_dispatch-only — no `schedule:`
trigger. Real LLM cost every run (Bedrock Haiku/GPT-4.1 for S1-old, Palmyra for
S1-from-atom) — do not wire this into any push/PR trigger without AA-287 Budgets alarms
in place first (Done as of this PR, but "done" != "wire this to run automatically" per the
issue's own explicit warning).

Fixture sourcing (AA-289 STEP 0 findings):
- S1-old: CIS_Golden_Tours_20_v1.xlsx, uploaded to
  s3://aa-cis-bronze-005097885195/fixtures/ (the repo's data/ dir is gitignored — the file
  is not committed, so a GitHub-Actions-triggered run has no local copy to read).
- S1-from-atom: the golden-tours fixture has no atom_id data at all, so it cannot exercise
  this pipeline. Uses the 4 real curated tours already verified in AA-306 instead
  (services/content_generation/s1_from_atom.py's own STEP 0/verify tours).
"""
import argparse
import asyncio
import io
import json
import os
import sys

import asyncpg
import boto3
import openpyxl
import structlog

logger = structlog.get_logger()

GOLDEN_TOURS_S3_BUCKET = "aa-cis-bronze-005097885195"
GOLDEN_TOURS_S3_KEY = "fixtures/CIS_Golden_Tours_20_v1.xlsx"
AWS_REGION = os.environ.get("AWS_REGION", "us-west-1")

# AA-289: absolute-point drop on a 0-10 scale. Same order of magnitude as graph.py's own
# MISSING_FIELD_CAP (4.0) / a single hard-block deduction — chosen so the gate fires on a
# real, single-issue-class regression (e.g. AA-231/AA-195's precedent bugs), not on the
# ordinary run-to-run noise of an LLM call.
S1_OLD_REGRESSION_THRESHOLD = 1.0

# S1-from-atom has no judge/quality_score (STEP 0 finding) — regression there is defined as a
# hard correctness signal instead of a soft numeric drift on a 4-tour sample: any tour that
# used to pass the grounding gate now failing it, or the gate pass rate dropping at all.
S1_FROM_ATOM_TOUR_IDS = [
    "c410a272-cac2-486d-9911-a5a73f5365d2",   # Classic Exploration, 21 atoms
    "cc2eca43-5422-4783-8a60-53cf0aef3003",   # Classic Exploration, 18 atoms
    "a3afaed4-4492-45d7-9a41-ff60314d4439",   # A sea of coral, Prayers of Uminchu fishermen, 6 atoms
    "296abe67-80e6-43c8-81af-2088b111d31f",   # Yaksa Trek BEST DEAL, 2 atoms (thin-trip edge case)
]


def _download_golden_tours() -> list[dict]:
    s3 = boto3.client("s3", region_name=AWS_REGION)
    obj = s3.get_object(Bucket=GOLDEN_TOURS_S3_BUCKET, Key=GOLDEN_TOURS_S3_KEY)
    wb = openpyxl.load_workbook(io.BytesIO(obj["Body"].read()), data_only=True)
    ws = wb["Golden Tours"]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    tours = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0] is None:
            continue
        rec = dict(zip(header, row))
        highlights = [h.strip() for h in (rec.get("highlights") or "").split("\n") if h.strip()]
        tours.append({
            "name":        rec.get("name") or "",
            "subtitle":    rec.get("subtitle") or "",
            "summary":     rec.get("summary") or "",
            "description": "",  # not present in the golden fixture — no equivalent field
            "highlights":  highlights,
            "itineraries": rec.get("itinerary_summary") or "",
            "country":     rec.get("country") or "",
            "duration":    rec.get("duration") or "",
            "price":       rec.get("price_usd") or "",
            "inclusions":  rec.get("inclusions") or "",
            "exclusions":  "",
            "_tour_id":    rec.get("tour_id"),  # golden-set id, not a real raw_tours UUID
        })
    return tours


async def run_s1_old_eval() -> dict:
    """Runs every Golden Tour through the REAL old-S1 LangGraph pipeline (generate -> validate
    -> llm_judge -> brand_audit -> flag_fix -> revalidate), no brand_rules (default AA voice —
    judge_node skips its GPT-4.1 call with no differentiation profile; brand_audit_node still
    runs its own GPT-4.1 call for legacy/no-profile brands, per its own docstring), seo={}
    (skip DataForSEO — an eval run scores writing quality, not live keyword data)."""
    from api.routers.v1_pipeline import _rewrite_tour

    tours = _download_golden_tours()
    results = []
    for idx, tour in enumerate(tours):
        result = await _rewrite_tour(
            tour, idx=idx, total=len(tours), brand_rules={}, seo={}, model_tier="haiku",
        )
        results.append(result)
        logger.info("eval_s1_old_tour_done", idx=idx, name=tour["name"],
                    quality_score=result.get("quality_score"), status=result.get("status"))

    scored = [r for r in results if r.get("status") == "success"]
    prompt_versions = {r.get("prompt_version") for r in scored if r.get("prompt_version")}
    if len(prompt_versions) > 1:
        logger.warning("eval_s1_old_prompt_version_mismatch", versions=list(prompt_versions))

    avg_score = round(sum(r["quality_score"] for r in scored) / len(scored), 3) if scored else None
    total_cost = round(sum(r.get("cost_usd", 0.0) for r in results), 4)
    return {
        "pipeline": "s1_old",
        "prompt_version": next(iter(prompt_versions), None) or "",
        "tour_count": len(tours),
        "avg_quality_score": avg_score,
        "cost_usd": total_cost,
        "details": {
            "scored_count": len(scored),
            "failed_count": len(results) - len(scored),
            "per_tour": [
                {"name": r.get("src_name"), "quality_score": r.get("quality_score"),
                 "status": r.get("status")}
                for r in results
            ],
        },
    }


async def run_s1_from_atom_eval(pool) -> dict:
    from services.content_generation.s1_from_atom import GroundingError, generate_s1_from_atom

    per_tour = []
    words_per_citation_values = []
    prompt_versions = set()
    total_cost_input_tokens = 0
    total_output_tokens = 0

    for tour_id in S1_FROM_ATOM_TOUR_IDS:
        async with pool.acquire() as conn:
            row = await conn.fetchrow(
                "SELECT src_name, country FROM silver_aa_internal.raw_tours WHERE tour_id = $1::uuid",
                tour_id,
            )
        tour = {"name": row["src_name"], "country": row["country"]} if row else {"name": "", "country": ""}
        try:
            result = await generate_s1_from_atom(tour_id, tour, pool, model_tier="palmyra")
            per_tour.append({
                "tour_id": tour_id, "name": tour["name"], "status": "passed",
                "words_per_citation": result["gate"]["words_per_citation"],
                "prompt_version": result["prompt_version"],
            })
            words_per_citation_values.append(result["gate"]["words_per_citation"])
            prompt_versions.add(result["prompt_version"])
            total_cost_input_tokens += result["input_tokens"]
            total_output_tokens += result["output_tokens"]
        except GroundingError as e:
            per_tour.append({
                "tour_id": tour_id, "name": tour["name"], "status": "gate_failed",
                "error": str(e), "prompt_version": e.prompt_version,
            })
            if e.prompt_version:
                prompt_versions.add(e.prompt_version)
        logger.info("eval_s1_from_atom_tour_done", tour_id=tour_id, status=per_tour[-1]["status"])

    if len(prompt_versions) > 1:
        logger.warning("eval_s1_from_atom_prompt_version_mismatch", versions=list(prompt_versions))

    pass_count = sum(1 for t in per_tour if t["status"] == "passed")
    avg_wpc = (
        round(sum(words_per_citation_values) / len(words_per_citation_values), 2)
        if words_per_citation_values else None
    )
    return {
        "pipeline": "s1_from_atom",
        "prompt_version": next(iter(prompt_versions), None) or "",
        "tour_count": len(S1_FROM_ATOM_TOUR_IDS),
        "avg_words_per_citation": avg_wpc,
        "gate_pass_count": pass_count,
        "gate_fail_count": len(S1_FROM_ATOM_TOUR_IDS) - pass_count,
        # No real $-rate for Palmyra confirmed yet (AA-289 STEP 0 didn't chase Bedrock's
        # Writer-model pricing page) — token counts only, not fabricated into a cost_usd number.
        "cost_usd": None,
        "details": {"per_tour": per_tour, "input_tokens": total_cost_input_tokens,
                     "output_tokens": total_output_tokens},
    }


async def _get_baseline(conn, pipeline: str, current_prompt_version: str):
    return await conn.fetchrow(
        """
        SELECT prompt_version, avg_quality_score, avg_words_per_citation, gate_pass_count
        FROM shared.prompt_eval_runs
        WHERE pipeline = $1 AND prompt_version != $2
        ORDER BY created_at DESC LIMIT 1
        """,
        pipeline, current_prompt_version,
    )


def _detect_regression(pipeline: str, current: dict, baseline) -> bool:
    if baseline is None:
        return False  # first-ever run for this pipeline — nothing to regress against
    if pipeline == "s1_old":
        if current["avg_quality_score"] is None or baseline["avg_quality_score"] is None:
            return False
        return (baseline["avg_quality_score"] - current["avg_quality_score"]) > S1_OLD_REGRESSION_THRESHOLD
    # s1_from_atom: any drop in gate pass rate vs baseline's pass rate on the same fixed 4-tour set
    if baseline["gate_pass_count"] is None:
        return False
    return current["gate_pass_count"] < baseline["gate_pass_count"]


async def _write_eval_run(conn, result: dict, baseline, regression: bool, triggered_by: str) -> None:
    await conn.execute(
        """
        INSERT INTO shared.prompt_eval_runs (
            pipeline, prompt_version, tour_count, avg_quality_score, avg_words_per_citation,
            gate_pass_count, gate_fail_count, cost_usd, regression_detected,
            baseline_prompt_version, baseline_avg_quality_score, details, triggered_by
        ) VALUES ($1, $2, $3, $4, $5, $6, $7, $8, $9, $10, $11, $12::jsonb, $13)
        """,
        result["pipeline"], result["prompt_version"], result["tour_count"],
        result.get("avg_quality_score"), result.get("avg_words_per_citation"),
        result.get("gate_pass_count"), result.get("gate_fail_count"), result.get("cost_usd"),
        regression, baseline["prompt_version"] if baseline else None,
        baseline["avg_quality_score"] if baseline else None,
        json.dumps(result.get("details", {}), default=str), triggered_by,
    )


async def run_eval(pipeline: str, triggered_by: str = "manual") -> dict:
    pool = await asyncpg.create_pool(os.environ["DATABASE_URL"], min_size=1, max_size=2)
    try:
        if pipeline == "s1_old":
            result = await run_s1_old_eval()
        elif pipeline == "s1_from_atom":
            result = await run_s1_from_atom_eval(pool)
        else:
            raise ValueError(f"Unknown pipeline: {pipeline!r}")

        async with pool.acquire() as conn:
            baseline = await _get_baseline(conn, pipeline, result["prompt_version"])
            regression = _detect_regression(pipeline, result, baseline)
            await _write_eval_run(conn, result, baseline, regression, triggered_by)

        result["baseline"] = dict(baseline) if baseline else None
        result["regression_detected"] = regression
        return result
    finally:
        await pool.close()


def main() -> int:
    parser = argparse.ArgumentParser(description="AA-289 on-demand prompt regression eval gate")
    parser.add_argument("--pipeline", choices=["s1_old", "s1_from_atom", "both"], required=True)
    parser.add_argument("--triggered-by", default="manual")
    args = parser.parse_args()

    pipelines = ["s1_old", "s1_from_atom"] if args.pipeline == "both" else [args.pipeline]
    any_regression = False
    for p in pipelines:
        result = asyncio.run(run_eval(p, triggered_by=args.triggered_by))
        print(json.dumps(result, indent=2, default=str))
        any_regression = any_regression or result["regression_detected"]

    return 1 if any_regression else 0


if __name__ == "__main__":
    sys.exit(main())
